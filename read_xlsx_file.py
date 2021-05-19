import openpyxl as op
import logging

logger = logging.getLogger(__name__)

excel_fname = "test1.xlsx" 	
def print_location_error():
    logger.error("File not found.\n")

def enter_filename():
    filename = input("Please enter file name.\n")
    return filename

def read_excel_file():
    try:
        open_file=op.load_workbook(filename=excel_fname, read_only=True)
        sheet=open_file.active
        sheet_list=open_file.sheetnames
        max_col = sheet.max_column
        max_row = sheet.max_row
        print("Active Sheet: ", sheet.title,'Row number: ', max_col,'Column number',max_row)
        for i in range (0,len(sheet_list)):
            sheet=open_file.worksheets[i]
            print(sheet.title)
            for row in sheet.iter_rows():
                for cell in row:
                    print(cell.value, end=" ")
                print()
        open_file.close()
    except FileNotFoundError:
        print_location_error()
        
read_excel_file()